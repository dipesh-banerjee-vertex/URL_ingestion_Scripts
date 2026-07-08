#!/usr/bin/env python3
"""Two-step catalog workflow.

Run 1: format source workbook to hardcoded _lodg columns.
Run 2: remove rows marked Invalid in a validation workbook.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook


# Hardcoded _lodg template columns (copied from _Lodg_CA format).
CANONICAL_COLUMNS: list[str] = [
    "tax_rule_id",
    "country",
    "state",
    "jur_name",
    "jur_id",
    "jur_type",
    "imposition_name",
    "imposition_type_name",
    "cat_id",
    "cat_name",
    "content_type",
    "rule_type",
    "url_type",
    "url",
    "skip_tool_running",
    "tag",
    "comment",
    "last_monitoring_date",
    "extract_text",
]


# Header aliases so Lodg_TN-style files still map correctly.
COLUMN_ALIASES: dict[str, list[str]] = {
    "tax_rule_id": ["taxruleid"],
    "jur_name": ["jurname", "jurisdiction_name"],
    "jur_id": ["jurid", "jurisdiction_id"],
    "jur_type": ["jurtype", "jurisdiction_type_id", "jurisdiction_type"],
    "imposition_name": ["impositionname"],
    "imposition_type_name": ["impositiontype", "imposition_type"],
    "cat_id": ["catid", "category_id"],
    "cat_name": ["catname", "category_name"],
    "rule_type": ["ruletype"],
}


@dataclass
class FormatResult:
    source_file: str
    source_sheet: str
    output_file: str
    source_rows: int
    formatted_rows: int
    columns_written: int


@dataclass
class ValidateResult:
    formatted_file: str
    formatted_sheet: str
    validation_file: str
    validation_sheet: str
    output_file: str
    source_rows: int
    invalid_url_count: int
    removed_rows: int
    final_rows: int


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_token(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def _normalize_url(value: object, case_insensitive: bool) -> str:
    text = _normalize_text(value)
    return text.lower() if case_insensitive else text


def _pick_sheet_name(
    all_sheets: Sequence[str],
    requested_name: str | None,
    preferred_names: Sequence[str],
) -> str:
    if requested_name:
        if requested_name not in all_sheets:
            raise ValueError(
                f"Sheet '{requested_name}' not found. Available sheets: {all_sheets}"
            )
        return requested_name

    for candidate in preferred_names:
        if candidate in all_sheets:
            return candidate

    return all_sheets[0]


def _read_headers(
    path: Path,
    sheet_name: str | None,
    preferred_names: Sequence[str],
) -> tuple[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet = _pick_sheet_name(wb.sheetnames, sheet_name, preferred_names)
        ws = wb[selected_sheet]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row is None:
            raise ValueError(f"Workbook has no header row: {path}")

        headers = [_normalize_text(v) for v in first_row]
        if not any(headers):
            raise ValueError(f"Header row is empty in {path} / {selected_sheet}")

        seen: set[str] = set()
        duplicates: list[str] = []
        for header in headers:
            token = _header_token(header)
            if not token:
                continue
            if token in seen:
                duplicates.append(header)
            seen.add(token)
        if duplicates:
            raise ValueError(
                f"Duplicate header(s) found in {path} / {selected_sheet}: {duplicates}"
            )

        return selected_sheet, headers
    finally:
        wb.close()


def _iter_data_rows(path: Path, sheet_name: str) -> Iterable[tuple[object, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield row
    finally:
        wb.close()


def _build_header_index_map(headers: Sequence[str]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for idx, header in enumerate(headers):
        token = _header_token(header)
        if token:
            index_map[token] = idx
    return index_map


def _candidate_tokens(canonical_column: str) -> list[str]:
    tokens = [_header_token(canonical_column)]
    for alias in COLUMN_ALIASES.get(canonical_column, []):
        alias_token = _header_token(alias)
        if alias_token and alias_token not in tokens:
            tokens.append(alias_token)
    return tokens


def _value_from_row(
    row: Sequence[object],
    index_map: dict[str, int],
    canonical_column: str,
) -> object:
    for token in _candidate_tokens(canonical_column):
        idx = index_map.get(token)
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _write_workbook(
    output_path: Path,
    sheet_name: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    wb.save(output_path)


def _write_audit(path: Path, payload: dict[str, object]) -> None:
    with path.open("w", encoding="utf-8") as file_handle:
        json.dump(payload, file_handle, indent=2)


def _default_formatted_name(source_path: Path) -> Path:
    return source_path.with_name(f"{source_path.stem}_formatted.xlsx")


def _default_validated_name(formatted_path: Path) -> Path:
    return formatted_path.with_name(f"{formatted_path.stem}_validated.xlsx")


def run_format(args: argparse.Namespace) -> int:
    source_path = Path(args.source)
    output_path = Path(args.output) if args.output else _default_formatted_name(source_path)

    source_sheet, source_headers = _read_headers(
        source_path,
        args.source_sheet,
        preferred_names=["Library"],
    )
    source_index_map = _build_header_index_map(source_headers)

    formatted_rows: list[list[object]] = []
    source_rows = 0

    for source_row in _iter_data_rows(source_path, source_sheet):
        source_rows += 1
        out_row: list[object] = []
        for column in CANONICAL_COLUMNS:
            if column == "extract_text":
                out_row.append(args.extract_text_value)
            else:
                out_row.append(_value_from_row(source_row, source_index_map, column))
        formatted_rows.append(out_row)

    result = FormatResult(
        source_file=str(source_path),
        source_sheet=source_sheet,
        output_file=str(output_path),
        source_rows=source_rows,
        formatted_rows=len(formatted_rows),
        columns_written=len(CANONICAL_COLUMNS),
    )

    if not args.dry_run:
        _write_workbook(output_path, source_sheet, CANONICAL_COLUMNS, formatted_rows)
        _write_audit(output_path.with_suffix(output_path.suffix + ".audit.json"), result.__dict__)

    print(json.dumps(result.__dict__, indent=2))
    return 0


def _find_column_index(headers: Sequence[str], column_name: str) -> int | None:
    wanted = _header_token(column_name)
    for idx, header in enumerate(headers):
        if _header_token(header) == wanted:
            return idx
    return None


def _build_invalid_url_set(
    validation_path: Path,
    validation_sheet: str | None,
    url_column: str,
    status_column: str,
    invalid_value: str,
    case_insensitive_url: bool,
) -> tuple[set[str], str]:
    sheet_name, headers = _read_headers(
        validation_path,
        validation_sheet,
        preferred_names=["Results", "Validation", "Library"],
    )

    url_idx = _find_column_index(headers, url_column)
    status_idx = _find_column_index(headers, status_column)
    if url_idx is None:
        raise ValueError(f"URL column '{url_column}' not found in validation workbook")
    if status_idx is None:
        raise ValueError(
            f"Validation status column '{status_column}' not found in validation workbook"
        )

    invalid_value_token = _header_token(invalid_value)
    invalid_urls: set[str] = set()

    for row in _iter_data_rows(validation_path, sheet_name):
        raw_status = "" if status_idx >= len(row) else _normalize_text(row[status_idx])
        if _header_token(raw_status) != invalid_value_token:
            continue

        raw_url = "" if url_idx >= len(row) else row[url_idx]
        normalized_url = _normalize_url(raw_url, case_insensitive_url)
        if normalized_url:
            invalid_urls.add(normalized_url)

    return invalid_urls, sheet_name


def run_validate(args: argparse.Namespace) -> int:
    formatted_path = Path(args.formatted)
    validation_path = Path(args.validation)
    output_path = Path(args.output) if args.output else _default_validated_name(formatted_path)

    source_sheet, source_headers = _read_headers(
        formatted_path,
        args.source_sheet,
        preferred_names=["Library"],
    )

    url_idx = _find_column_index(source_headers, args.url_column)
    if url_idx is None:
        raise ValueError(
            f"URL column '{args.url_column}' not found in formatted workbook {formatted_path}"
        )

    invalid_urls, validation_sheet = _build_invalid_url_set(
        validation_path=validation_path,
        validation_sheet=args.validation_sheet,
        url_column=args.url_column,
        status_column=args.validation_status_column,
        invalid_value=args.invalid_value,
        case_insensitive_url=args.url_case_insensitive,
    )

    kept_rows: list[list[object]] = []
    source_rows = 0
    removed_rows = 0

    for row in _iter_data_rows(formatted_path, source_sheet):
        source_rows += 1
        row_url = "" if url_idx >= len(row) else row[url_idx]
        normalized_url = _normalize_url(row_url, args.url_case_insensitive)

        if normalized_url and normalized_url in invalid_urls:
            removed_rows += 1
            continue

        kept_rows.append(list(row))

    result = ValidateResult(
        formatted_file=str(formatted_path),
        formatted_sheet=source_sheet,
        validation_file=str(validation_path),
        validation_sheet=validation_sheet,
        output_file=str(output_path),
        source_rows=source_rows,
        invalid_url_count=len(invalid_urls),
        removed_rows=removed_rows,
        final_rows=len(kept_rows),
    )

    if not args.dry_run:
        _write_workbook(output_path, source_sheet, source_headers, kept_rows)
        _write_audit(output_path.with_suffix(output_path.suffix + ".audit.json"), result.__dict__)

    print(json.dumps(result.__dict__, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Two-step Lodg ingestion workflow")
    subparsers = parser.add_subparsers(dest="command", required=True)

    format_parser = subparsers.add_parser(
        "format",
        help="Run step 1: keep hardcoded _lodg columns and force extract_text=yes",
    )
    format_parser.add_argument("--source", required=True, help="Input workbook in Lodg_TN style")
    format_parser.add_argument("--output", help="Output workbook path (default: <source>_formatted.xlsx)")
    format_parser.add_argument("--source-sheet", help="Source sheet name (default: Library, else first)")
    format_parser.add_argument(
        "--extract-text-value",
        default="yes",
        help="Value to write in extract_text column for all rows",
    )
    format_parser.add_argument("--dry-run", action="store_true", help="Preview counts without writing files")

    validate_parser = subparsers.add_parser(
        "validate",
        help="Run step 2: remove rows marked Invalid in validation workbook",
    )
    validate_parser.add_argument("--formatted", required=True, help="Formatted workbook from step 1")
    validate_parser.add_argument("--validation", required=True, help="Validation workbook")
    validate_parser.add_argument(
        "--output",
        help="Output workbook path (default: <formatted>_validated.xlsx)",
    )
    validate_parser.add_argument("--source-sheet", help="Formatted workbook sheet name")
    validate_parser.add_argument("--validation-sheet", help="Validation workbook sheet name")
    validate_parser.add_argument("--url-column", default="url", help="URL column name")
    validate_parser.add_argument(
        "--validation-status-column",
        default="validation_status",
        help="Status column in validation workbook",
    )
    validate_parser.add_argument(
        "--invalid-value",
        default="Invalid",
        help="Status value marking invalid rows",
    )
    validate_parser.add_argument(
        "--url-case-insensitive",
        action="store_true",
        help="Compare URLs case-insensitively",
    )
    validate_parser.add_argument("--dry-run", action="store_true", help="Preview counts without writing files")

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if args.command == "format":
            return run_format(args)
        if args.command == "validate":
            return run_validate(args)
        parser.error("Unknown command")
        return 2
    except Exception as exc:  # noqa: BLE001 - CLI should print friendly errors.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
